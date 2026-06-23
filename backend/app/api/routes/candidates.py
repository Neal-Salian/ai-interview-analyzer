import os
import io
import csv
import re
import uuid
import datetime
import logging
import shutil
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional

from app.db.database import get_db
from app.db.models import Candidate, UserRole, Session as InterviewSession, Job
from app.api.deps import get_current_user, require_recruiter

router = APIRouter()
logger = logging.getLogger(__name__)

# ── Max file size for bulk import (5 MB) ──────────────────────────────────────
MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024

# ── Email regex (same standard used across the project) ───────────────────────
EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$")


class CandidateCreate(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None
    notes: Optional[str] = None
    job_id: Optional[str] = None


class CandidateUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None


class BulkCandidateEntry(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None
    notes: Optional[str] = None
    job_id: Optional[str] = None


class BulkImportRequest(BaseModel):
    candidates: list[BulkCandidateEntry]
    job_id: Optional[str] = None  # fallback job_id for all candidates


# ── Shared creation helper ────────────────────────────────────────────────────
# Both single-create and bulk-import use this to guarantee identical behavior.

def _create_candidate_with_session(
    db: Session,
    recruiter_id: uuid.UUID,
    name: str,
    email: str,
    phone: Optional[str] = None,
    notes: Optional[str] = None,
    job_id: Optional[str] = None,
) -> tuple:
    """
    Creates a Candidate + Draft Session in the given db session (no commit).
    Returns (candidate, draft_session).
    Caller is responsible for commit/rollback.
    """
    candidate = Candidate(
        id=uuid.uuid4(),
        recruiter_id=recruiter_id,
        name=name,
        email=email,
        phone=phone,
        notes=notes,
        status="Draft",
        created_at=datetime.datetime.utcnow()
    )
    db.add(candidate)

    # Auto-create draft session
    session_id = uuid.uuid4()
    job_uuid = None
    if job_id:
        try:
            job_uuid = uuid.UUID(job_id)
        except ValueError:
            pass

    draft_session = InterviewSession(
        id=session_id,
        candidate_id=candidate.id,
        job_id=job_uuid,
        recruiter_id=recruiter_id,
        status="draft",
        scheduled_at=None,
        started_at=None
    )
    db.add(draft_session)

    return candidate, draft_session


# ── Single candidate create (unchanged behavior) ─────────────────────────────

@router.post("/candidates")
def create_candidate(
    payload: CandidateCreate,
    db: Session = Depends(get_db),
    current_user=Depends(require_recruiter)
):
    existing = db.query(Candidate).filter(
        Candidate.email == payload.email
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    candidate, draft_session = _create_candidate_with_session(
        db=db,
        recruiter_id=current_user.id,
        name=payload.name,
        email=payload.email,
        phone=payload.phone,
        notes=payload.notes,
        job_id=payload.job_id,
    )

    db.commit()
    db.refresh(candidate)
    logger.info(f"[candidates] created {candidate.email} by recruiter {current_user.id} and draft session {draft_session.id}")
    return {
        "candidate_id": str(candidate.id),
        "name": candidate.name,
        "session_id": str(draft_session.id)
    }


# ── Candidate listing ────────────────────────────────────────────────────────

@router.get("/candidates")
def list_candidates(
    db: Session = Depends(get_db),
    current_user=Depends(require_recruiter)
):
    query = db.query(Candidate)
    if current_user.role != UserRole.ADMIN:
        query = query.filter(Candidate.recruiter_id == current_user.id)
        
    candidates = query.order_by(Candidate.created_at.desc()).all()
    
    result = []
    for c in candidates:
        applied_jobs = [s.job.title for s in c.sessions if s.job]
        
        derived_status = "Draft"
        if any(s.status in ["scheduled", "active", "processing"] for s in c.sessions):
            derived_status = "Scheduled"
        elif any(s.status == "completed" for s in c.sessions):
            derived_status = "Completed"

        result.append({
            "id": str(c.id),
            "name": c.name,
            "email": c.email,
            "phone": c.phone,
            "status": derived_status,
            "applied_jobs": list(set(applied_jobs)),
            "created_at": c.created_at.isoformat()
        })
    return result


# ── Get single candidate ─────────────────────────────────────────────────────

@router.get("/candidates/{candidate_id}")
def get_candidate(
    candidate_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_recruiter)
):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
        
    if current_user.role != UserRole.ADMIN and candidate.recruiter_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to view this candidate")

    applied_jobs = []
    session_history = []
    
    for s in candidate.sessions:
        if s.job:
            applied_jobs.append({"id": str(s.job.id), "title": s.job.title})
        session_history.append({
            "session_id": str(s.id),
            "status": s.status,
            "scheduled_at": s.scheduled_at,
            "started_at": s.started_at,
            "ended_at": s.ended_at,
            "job": s.job.title if s.job else None,
            "job_id": str(s.job.id) if s.job else None
        })

    derived_status = "Draft"
    if any(s.status in ["scheduled", "active", "processing"] for s in candidate.sessions):
        derived_status = "Scheduled"
    elif any(s.status == "completed" for s in candidate.sessions):
        derived_status = "Completed"

    # Unique jobs
    unique_jobs = {j["id"]: j for j in applied_jobs}.values()

    return {
        "id": str(candidate.id),
        "name": candidate.name,
        "email": candidate.email,
        "phone": candidate.phone,
        "notes": candidate.notes,
        "status": derived_status,
        "resume_url": candidate.resume_url,
        "created_at": candidate.created_at.isoformat(),
        "applied_jobs": list(unique_jobs),
        "session_history": session_history
    }


# ── Update candidate ─────────────────────────────────────────────────────────

@router.patch("/candidates/{candidate_id}")
def update_candidate(
    candidate_id: str,
    payload: CandidateUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(require_recruiter)
):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
        
    if current_user.role != UserRole.ADMIN and candidate.recruiter_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this candidate")

    if payload.name is not None:
        candidate.name = payload.name
    if payload.email is not None:
        # Check uniqueness
        if payload.email != candidate.email:
            existing = db.query(Candidate).filter(Candidate.email == payload.email).first()
            if existing:
                raise HTTPException(status_code=400, detail="Email already used")
        candidate.email = payload.email
    if payload.phone is not None:
        candidate.phone = payload.phone
    if payload.notes is not None:
        candidate.notes = payload.notes

    db.commit()
    db.refresh(candidate)
    
    return {"status": "success", "candidate_id": str(candidate.id)}


# ── Resume upload ─────────────────────────────────────────────────────────────

@router.post("/candidates/{candidate_id}/resume")
def upload_resume(
    candidate_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_recruiter)
):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
        
    if current_user.role != UserRole.ADMIN and candidate.recruiter_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this candidate")

    allowed_types = ["application/pdf", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"]
    if file.content_type not in allowed_types:
        if not (file.filename.endswith(".pdf") or file.filename.endswith(".docx")):
            raise HTTPException(status_code=400, detail="Only PDF and DOCX files are allowed")

    upload_dir = "uploads"
    os.makedirs(upload_dir, exist_ok=True)
    
    ext = ".pdf" if file.filename.endswith(".pdf") else ".docx"
    safe_filename = f"resume_{candidate.id}{ext}"
    file_path = os.path.join(upload_dir, safe_filename)

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    candidate.resume_url = f"/api/uploads/{safe_filename}"
    db.commit()
    
    return {"status": "success", "resume_url": candidate.resume_url}


# ══════════════════════════════════════════════════════════════════════════════
# BULK IMPORT ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

def _parse_csv_file(content: bytes) -> list[dict]:
    """Parse CSV bytes into list of row dicts. Normalises header names."""
    text = content.decode("utf-8-sig")  # handles BOM
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        normalised = {}
        for key, value in row.items():
            if key is not None:
                normalised[key.strip().lower()] = (value or "").strip()
        rows.append(normalised)
    return rows


def _parse_xlsx_file(content: bytes) -> list[dict]:
    """Parse XLSX bytes into list of row dicts. Normalises header names."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    
    # First row = headers
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return []
    
    headers = [(str(h).strip().lower() if h else f"col_{i}") for i, h in enumerate(raw_headers)]
    
    rows = []
    for row_values in rows_iter:
        row_dict = {}
        for i, val in enumerate(row_values):
            if i < len(headers):
                row_dict[headers[i]] = str(val).strip() if val is not None else ""
        # Skip completely empty rows
        if any(v for v in row_dict.values()):
            rows.append(row_dict)
    
    wb.close()
    return rows


def _normalise_column_name(col: str) -> str:
    """Map common column name variations to canonical field names."""
    col = col.strip().lower()
    mappings = {
        "name": "name", "full name": "name", "full_name": "name", "candidate name": "name", "candidate_name": "name",
        "email": "email", "email address": "email", "email_address": "email", "e-mail": "email",
        "phone": "phone", "phone number": "phone", "phone_number": "phone", "mobile": "phone", "contact": "phone",
        "notes": "notes", "note": "notes", "comments": "notes", "comment": "notes",
        "job": "job", "job title": "job", "job_title": "job", "job id": "job", "job_id": "job", "position": "job", "role": "job",
    }
    return mappings.get(col, col)


def _validate_rows(rows: list[dict], db: Session, recruiter_id: uuid.UUID, fallback_job_id: Optional[str] = None) -> dict:
    """
    Validate parsed rows and return {valid, invalid, duplicates} lists.
    Does NOT create any records.
    """
    valid = []
    invalid = []
    duplicates = []

    # Normalise column names in rows
    normalised_rows = []
    for row in rows:
        normalised = {}
        for key, value in row.items():
            canon = _normalise_column_name(key)
            normalised[canon] = value
        normalised_rows.append(normalised)

    # Pre-fetch existing candidate emails for this recruiter (batch query)
    existing_emails_query = db.query(Candidate.email).all()
    existing_emails = {e[0].lower() for e in existing_emails_query}

    # Load recruiter's jobs for matching
    job_query = db.query(Job).filter(Job.is_archived == False)
    if recruiter_id:
        # Include jobs owned by recruiter or admin-visible jobs
        job_query = job_query  # Jobs are filtered by recruiter in the route, but for validation we need all accessible
    all_jobs = job_query.all()
    
    # Build lookup maps
    job_by_title = {j.title.lower(): j for j in all_jobs}
    job_by_id = {}
    for j in all_jobs:
        job_by_id[str(j.id)] = j
        job_by_id[str(j.id).lower()] = j

    # Track emails seen within this file to detect intra-file duplicates
    seen_emails = set()

    for idx, row in enumerate(normalised_rows):
        row_num = idx + 2  # +2 because row 1 is headers, data starts at row 2
        name = row.get("name", "").strip()
        email = row.get("email", "").strip().lower()
        phone = row.get("phone", "").strip()
        notes = row.get("notes", "").strip()
        job_value = row.get("job", "").strip()

        errors = []

        # Required field validation
        if not name:
            errors.append("Name is required")
        if not email:
            errors.append("Email is required")
        elif not EMAIL_REGEX.match(email):
            errors.append("Invalid email format")

        # Intra-file duplicate check
        if email and email in seen_emails:
            errors.append("Duplicate email within file")

        if errors:
            invalid.append({
                "row": row_num,
                "name": name,
                "email": email,
                "phone": phone,
                "notes": notes,
                "errors": errors
            })
            if email:
                seen_emails.add(email)
            continue

        # Database duplicate check
        if email in existing_emails:
            duplicates.append({
                "row": row_num,
                "name": name,
                "email": email,
                "phone": phone,
                "notes": notes,
                "reason": "Email already exists in database"
            })
            seen_emails.add(email)
            continue

        # Job resolution
        resolved_job_id = None
        resolved_job_title = None
        
        if job_value:
            # Try matching by UUID first
            matched_job = job_by_id.get(job_value) or job_by_id.get(job_value.lower())
            if not matched_job:
                # Try matching by title (case-insensitive)
                matched_job = job_by_title.get(job_value.lower())
            
            if matched_job:
                resolved_job_id = str(matched_job.id)
                resolved_job_title = matched_job.title
            else:
                errors.append(f"Job not found: '{job_value}'")
                invalid.append({
                    "row": row_num,
                    "name": name,
                    "email": email,
                    "phone": phone,
                    "notes": notes,
                    "errors": errors
                })
                seen_emails.add(email)
                continue
        elif fallback_job_id:
            matched_job = job_by_id.get(fallback_job_id) or job_by_id.get(fallback_job_id.lower())
            if matched_job:
                resolved_job_id = str(matched_job.id)
                resolved_job_title = matched_job.title

        seen_emails.add(email)
        valid.append({
            "row": row_num,
            "name": name,
            "email": email,
            "phone": phone if phone else None,
            "notes": notes if notes else None,
            "job_id": resolved_job_id,
            "job_title": resolved_job_title
        })

    return {"valid": valid, "invalid": invalid, "duplicates": duplicates}


@router.post("/candidates/import/validate")
async def validate_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_recruiter)
):
    """
    Validates a CSV/XLSX file for bulk candidate import.
    Returns categorised rows (valid/invalid/duplicate) without creating records.
    """
    # File type validation
    filename = file.filename or ""
    if not (filename.lower().endswith(".csv") or filename.lower().endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Only CSV and XLSX files are supported")

    # Read file content
    content = await file.read()

    # File size validation
    if len(content) > MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"File size exceeds {MAX_IMPORT_FILE_SIZE // (1024*1024)}MB limit")

    # Parse file
    try:
        if filename.lower().endswith(".csv"):
            rows = _parse_csv_file(content)
        else:
            rows = _parse_xlsx_file(content)
    except Exception as e:
        logger.error(f"[bulk-import] Failed to parse file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows")

    # Check required columns exist
    first_row_keys = {_normalise_column_name(k) for k in rows[0].keys()}
    if "name" not in first_row_keys:
        raise HTTPException(status_code=400, detail="Missing required column: Name")
    if "email" not in first_row_keys:
        raise HTTPException(status_code=400, detail="Missing required column: Email")

    # Validate all rows
    result = _validate_rows(rows, db, current_user.id)
    
    logger.info(
        f"[bulk-import] Validation for recruiter {current_user.id}: "
        f"{len(result['valid'])} valid, {len(result['invalid'])} invalid, {len(result['duplicates'])} duplicates"
    )
    
    return result


@router.post("/candidates/import")
def import_candidates(
    payload: BulkImportRequest,
    db: Session = Depends(get_db),
    current_user=Depends(require_recruiter)
):
    """
    Creates candidates from validated data. Skips duplicates.
    Uses the same creation logic as single-candidate creation.
    """
    created = 0
    skipped = 0
    failed = 0
    results = []

    for entry in payload.candidates:
        email = entry.email.strip().lower()
        
        # Check for existing candidate (skip duplicates)
        existing = db.query(Candidate).filter(Candidate.email == email).first()
        if existing:
            skipped += 1
            results.append({
                "email": email,
                "name": entry.name,
                "status": "skipped",
                "reason": "Email already exists"
            })
            continue

        # Resolve job_id: per-candidate job_id takes priority, then fallback
        effective_job_id = entry.job_id or payload.job_id

        try:
            candidate, draft_session = _create_candidate_with_session(
                db=db,
                recruiter_id=current_user.id,
                name=entry.name.strip(),
                email=email,
                phone=entry.phone.strip() if entry.phone else None,
                notes=entry.notes.strip() if entry.notes else None,
                job_id=effective_job_id,
            )
            created += 1
            results.append({
                "email": email,
                "name": entry.name,
                "status": "created",
                "candidate_id": str(candidate.id)
            })
        except Exception as e:
            logger.error(f"[bulk-import] Failed to create candidate {email}: {e}")
            failed += 1
            results.append({
                "email": email,
                "name": entry.name,
                "status": "failed",
                "reason": str(e)
            })

    # Single commit for all successful creates
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"[bulk-import] Transaction failed: {e}")
        raise HTTPException(status_code=500, detail="Import transaction failed. No candidates were created.")

    logger.info(
        f"[bulk-import] Import by recruiter {current_user.id}: "
        f"{created} created, {skipped} skipped, {failed} failed"
    )

    return {
        "created": created,
        "skipped": skipped,
        "failed": failed,
        "results": results
    }